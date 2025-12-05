import reflex as rx
from typing import Dict, List, Any, Optional
from .data import DataState
import pandas as pd
import io
import os

try:
    from src.reporting import ReportGenerator
    # We will simulate the analysis part if deep inspection is needed or duplicate logic
    BACKEND_AVAILABLE = True
except ImportError:
    BACKEND_AVAILABLE = False
    print("Warning: Reporting backend modules not found.")

class ReportingState(DataState):
    """State for Automated Reporting."""
    
    # Uploads
    template_filename: str = ""
    template_file_content: Optional[bytes] = None
    template_structure: Dict[str, Any] = {}
    
    # Mapping
    mapping_config: Dict[str, Any] = {}
    
    # Generation
    generated_report_path: str = ""
    is_generating: bool = False
    
    # UI Toggles
    active_step: int = 0 # 0: Upload, 1: Mapping, 2: Download

    async def handle_template_upload(self, files: List[rx.UploadFile]):
        """Handle template file upload."""
        if not files:
            return
            
        file = files[0]
        self.template_filename = file.filename
        content = await file.read()
        self.template_file_content = content
        
        # Analyze structure (MVP: Mock or simple)
        # Real impl would use openpyxl like streamlit code
        self.analyze_template_structure(file.filename, content)
        
        # Move to next step if data is also ready
        if self.template_structure:
             self.active_step = 1

    def analyze_template_structure(self, filename: str, content: bytes):
        """Analyze template to find placeholders."""
        # Simplified logic for MVP - assumes XLSX with {{placeholders}}
        if filename.endswith(".xlsx"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content))
                placeholders = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if isinstance(cell, str) and "{{" in cell and "}}" in cell:
                                # Extract content between braces
                                import re
                                matches = re.findall(r"\{\{(.*?)\}\}", cell)
                                for m in matches:
                                    placeholders.append(m.strip())
                
                # Remove duplicates
                placeholders = list(set(placeholders))
                
                self.template_structure = {
                    "type": "xlsx",
                    "placeholders": placeholders,
                    "sheet_count": len(wb.sheetnames)
                }
            except Exception as e:
                print(f"Error analyzing template: {e}")
                self.template_structure = {"error": str(e)}

    def set_mapping(self, placeholder: str, column: str):
        """Set mapping for a placeholder."""
        self.mapping_config[placeholder] = column

    def generate_report(self):
        """Generate the final report."""
        if not self.template_file_content:
             return rx.window_alert("No template uploaded.")
             
        self.is_generating = True
        yield
        
        try:
            # MVP Generation Logic
            # In a real app we'd call src.reporting.ReportGenerator
            
            # Simulated delay
            import time
            time.sleep(1)
            
            self.active_step = 2
            self.generated_report_path = "report_generated.xlsx" # Mock
            
        except Exception as e:
            return rx.window_alert(f"Generation Error: {e}")
        finally:
            self.is_generating = False
            
    def reset(self):
        self.template_filename = ""
        self.template_file_content = None
        self.template_structure = {}
        self.mapping_config = {}
        self.generated_report_path = ""
        self.active_step = 0
